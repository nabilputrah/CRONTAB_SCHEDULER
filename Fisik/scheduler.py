from openpyxl import Workbook
import subprocess

# Buat file Excel
wb = Workbook()
ws = wb.active
ws.title = "Jadwal Script"

# Dapatkan jadwal crontab untuk pengguna root
command = "crontab -l -u root"
result = subprocess.getoutput(command)
lines = result.split('\n')

print("Total Cron yang ada =",len(lines))

row = 1

for line in lines:
    # Ambil nama script dari path
    script_name = line.split()[-1]

    # Ambil menit dan jam dari jadwal crontab
    menit, jam = line.split()[:2]

    # Masukkan informasi ke dalam file Excel
    ws.cell(row=row, column=1, value="Nama = " + script_name)
    ws.cell(row=row+1, column=1, value="Menit = " + menit)
    ws.cell(row=row+2, column=1, value="Jam = " + jam)

    row += 4

# Simpan file Excel
wb.save("jadwal_script.xlsx")
print("File Excel jadwal_script.xlsx telah dibuat.")
